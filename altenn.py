from cProfile import *
from msilib.schema import Error
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
from tkinter import font
from click import command
from colorama import Cursor
from kivy.uix.scrollview import ScrollView
from openpyxl import load_workbook
from numpy import size 
import openpyxl
import pandas as pd
import os
from tkinter import*
from tkinter import filedialog
from openpyxl import *
from tkinter.messagebox import showinfo
from tkinter.messagebox import showinfo
from tkcalendar import Calendar
from tkcalendar import DateEntry
from PIL import ImageTk, Image
import tkinter.font as tkFont

wb_path=(r'C:\Users\Administrateur\Downloads\Classeur3.xlsx')
wb=load_workbook(wb_path)
employee_list=pd.read_excel(wb_path)
sheet_obj=wb.active

class studentForm:
    def __init__(self,root) :
        
        self.root = root
        
        self.root.withdraw()
        self.root.title("CRUD")
        x = (self.root.winfo_screenwidth() - 
             self.root.winfo_reqwidth() )/ 15
        y = (self.root.winfo_screenheight()-self.root.winfo_reqheight())/35
        self.root.geometry("1633x955+%d+%d" % (x,y))

        #self.root_ku= Button(self.root,  text = 'INSERT' , width=24 , height=2 ,bd=2 ,relief=FLAT , bg="#fc5c01" ,  fg="white", font=("roboto sans-serif",13, "bold"))
        #self.root_ku.pack(side=RIGHT, fill=X)
          
        def save(): # Récupération des données

            
            ID= self.employe_entry.get()
            Manager=self.comborr.get()
            Company=self.Company_entry.get()
            City= self.Citycombo.get()
            Department= self.Department_entry.get()
            Type_de_contract=self.combof.get()
            Last_project_name=self.project_entry.get()
            Last_position=self.position_entry.get()
            Hiring_date=self.dates_entry.get()
            Level=self.combo_lev.get()
            PCI=self.pci_entry.get()
            Specific_position=self.Specific_position_entry.get()
            Diplome=self.bac_lev.get()
            Performance=self.Performance_entry.get()
            Potentiel=self.comboPotential.get()
            Studies=self.Studies_entry.get()
            Scoring=self.comboScoring.get()
            Principal_competency=self.Principal_competency_entry.get()
            Secondary_competency=self.Secondary_competency_entry.get()
            Hiring_date_SISE=self.datesSISE_entry.get()
            M_F=self.comboMF.get()
            E_T=self.comboET.get()
            Leaving_reasons=self.Leaving_Reasons_entry.get()
            Absenteeism=self.Absenteeism_Reasons_entry.get()
            Previous_experience=self.Previous_exp_entry .get()
            Mission=self.Mission_entry.get()
            Date_CDI=self.datesCDI_entry.get()
            Exit_date=self.datesExit_Date_entry.get()
            experience_alten=self.Expérience_ALTEN_entry.get()
            experience_totale=self.Expérience_totale_entry.get()
            monthly_salary=self.Monthly_salary_local.get()
            prenom=self.Prenom_entry.get()
            nom=self.Nom_entry.get()
            
            wb=Workbook()
            ws=wb.active
            
            ws['A1']="ID"
            ws['B1']="Manager"
            ws['C1']="Company"
            ws['D1']="City"  
            ws['E1']="Department"
            ws['F1']="Type_de_contract"
            ws['G1']="Last_project_name"
            ws['H1']="Last_position"
            ws['I1']="Hiring_date"
            ws['J1']="Level"
            ws['K1']="PCI"
            ws['L1']="Specific_Position"
            ws['M1']="Diplôme"
            ws['N1']="Performance"
            ws['O1']="Potentiel"
            ws['Q1']="Studies"
            ws['R1']="Scoring"
            ws['R1']="Principal_competency"    
            ws['S1']="Secondary_competency"    
            ws['T1']="Hiring_date_SISE "
            ws['U1']="M/F"
            ws['V1']="E/T"
            ws['W1']="Leaving"
            ws['X1']="Abssenteeism"
            ws['Y1']="Previous experience"
            ws['Z1']="Mission"
            ws['AA1']="Date CDI"
            ws['AB1']=" Exit date"
            ws['AC1']="Experience alten"
            ws['AD1']="experience totale"
            ws['AE1']="monthly salary"
            ws['AF1']="Prenom"
            ws['AG1']="Nom"
            
            ws['A2']=ID
            ws['B2']=Manager
            ws['C2']=Company
            ws['D2']=City
            ws['E2']=Department
            ws['F2']=Type_de_contract
            ws['G2']=Last_project_name
            ws['H2']=Last_position
            ws['I2']=Hiring_date
            ws['J2']=Level
            ws['K2']=PCI
            ws['L2']=Specific_position
            ws['M2']=Diplome
            ws['N2']=Performance
            ws['O2']=Potentiel
            ws['Q2']=Studies
            ws['R2']=Scoring
            ws['R2']=Principal_competency
            ws['S2']=Secondary_competency
            ws['T2']=Hiring_date_SISE 
            ws['U2']=M_F
            ws['V2']=E_T
            ws['W2']=Leaving_reasons
            ws['X2']=Absenteeism
            ws['Y2']=Previous_experience
            ws['Z2']=Mission
            ws['AA2']=Date_CDI
            ws['AB2']= Exit_date
            ws['AC2']=experience_alten
            ws['AD2']=experience_totale
            ws['AE2']=monthly_salary
            ws['AF2']=prenom
            ws['AG2']=nom
            
            
            wb.save(r'C:\Users\Administrateur\Downloads\Classeur11.xlsx')
            showinfo("Saved", "Votre saisie a été enregistré !")
            file2=pd.read_excel(r'C:\Users\Administrateur\Downloads\Classeur11.xlsx')
            file1=pd.read_excel(r'C:\Users\Administrateur\Downloads\Classeur3.xlsx')
        
            all=[file1,file2]
            append=pd.concat(all)
            append.to_excel(r'C:\Users\Administrateur\Downloads\Classeur3.xlsx', index=False)
            self.data_list.insert('', '0', values=(self.employe_entry.get(), self.comborr.get(), self.Company_entry.get(),
                                                  self.Citycombo.get(),self.Department_entry.get(),
                                                  self.combof.get(),self.project_entry.get(), self.position_entry.get(),
                                                  self.dates_entry.get_date(),self.combo_lev.get(),
                                                  self.pci_entry.get(), self.Specific_position_entry.get(),
                                                  self.bac_lev.get(),self.Performance_entry.get(),self.comboPotential.get(),
                                                  self.Studies_entry.get(),self.comboScoring.get(),self.Principal_competency_entry.get(),
                                                  self.Secondary_competency_entry.get(), self.datesSISE_entry.get_date(),self.comboMF.get(),
                                                  self.comboET.get(), self.Leaving_Reasons_entry.get(),self.Absenteeism_Reasons_entry.get(),
                                                  self.Previous_exp_entry.get(),self.Mission_entry.get(),self.datesCDI_entry.get_date(),self.datesExit_Date_entry.get_date(),
                                                  self.Expérience_ALTEN_entry.get(), self.Expérience_totale_entry.get(), self.Monthly_salary_local_entry.get(),self.Prenom_entry.get(),self.Nom_entry.get()))
            
            
           
              
        #titre de la fenetre
        
        fontstyle=tkFont.Font(family="Lucida Grande", size=30)
        self.lbl_title = Label(
            self.root, text = "Informations des employés" ,padx=2,pady=14  , font=fontstyle,  fg="white" , 
            relief = RIDGE, bg="#225F85" )
        self.lbl_title.pack(side = TOP, fill= X)
       
        
        self.txt_frame = Frame(
            self.root , bd= 4, relief = RIDGE , bg="#CBDEF1")
        self.txt_frame.place(x=0,y=76, width=1638, height=700)
        
        #table
        self.detail_frame = Frame(
            self.root , bd= 4, relief = RIDGE , bg="#CBDEF1")
        self.detail_frame.place(x=0,y=470, width=1638, height=700)
        
        #Buttons
        self.btn_frame = Frame(
            self.root , bd= 4, relief = RIDGE , bg="#225F85")
        self.btn_frame.place(x=0,y=377, width=1638, height=93)
        
        #save
        self.save_btn = Button(self.btn_frame ,  text = 'Insérer' , command= save,width=24 , height=2 ,bd=2 ,relief=FLAT , bg="#E3D919" ,  fg="black",
                               font=("roboto sans-serif",13, "bold"))
        self.save_btn.grid(row = 1, column=3 , pady = 20 , padx= 10)
        
        #update
        self.update_btn = Button(self.btn_frame ,  text = 'Modifier' , width=24 , height=2 ,bd=2 ,relief=FLAT , bg="#E3D919" ,  fg="black",
                               font=("roboto sans-serif",13, "bold"), command= self.Update)
        self.update_btn.grid(row = 1, column=4 , pady = 20 , padx= 10)
        
        #select
        self.select_btn = Button(self.btn_frame ,  text = 'Selectionner' , width=24 , height=2 ,bd=2 ,relief=FLAT , bg="#E3D919" ,  fg="black",
                               font=("roboto sans-serif",13, "bold"),command=self.select_record)
        self.select_btn.grid(row = 1, column=7 , pady = 20 , padx= 10)
       
       #delete
        self.delete_btn = Button(self.btn_frame ,  text = 'Supprimer' , width=24 , height=2 ,bd=2 ,relief=FLAT , bg="#E3D919" ,  fg="black",
                               font=("roboto sans-serif",13, "bold"),command= self.delete_data)
        self.delete_btn.grid(row = 1, column=5 , pady = 20 , padx= 10)

        
        #clear
        self.clear_btn = Button(self.btn_frame ,  text = 'Effacer' , width=24 , height=2 ,bd=2 ,relief=FLAT , bg="#E3D919" ,  fg="black",
                               font=("roboto sans-serif",13, "bold"), command= self.clear)
        self.clear_btn.grid(row = 1, column=6 , pady = 20 , padx= 10)
        
        #space area
        self.space_label = Label(self.txt_frame, text="" , font=("", 10), bg="#CBDEF1", fg="#3299CC").grid(row=0, column=2,sticky=W)
        
        # employe's ID
        self.employe = StringVar()
        self.employe_label = Label(self.txt_frame, text="ID", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=1, column=0, sticky=W, padx=10)
        self.employe_entry = tk.Entry(self.txt_frame, textvariable=self.employe, width=17, bd=3, font=("bold",12))
        self.employe_entry.grid(row=1,column=1) 

        # Manager name
        self.opts = ('XXXXXXXXXX', 'YYYYYYYYYY')
        self.Manager = StringVar()
        self.Manager_label = Label(self.txt_frame, text="Manager", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=1, column=3, sticky=W, padx=5)
        self.comborr = ttk.Combobox(self.txt_frame, textvariable=self.Manager, values=self.opts,  width=16,  font=("bold",12))
        self.comborr.grid(row=1,column=4) 
        
        #Company
        self.Company = StringVar()
        self.Company_Label = Label(self.txt_frame, text="Entreprise", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=1, column=5, sticky=W, padx=5)
        self.Company_entry = tk.Entry(self.txt_frame, textvariable=self.Company, width=17, bd=3, font=("bold",12))
        self.Company_entry.grid(row=1,column=6)
        
        # City name
        self.Cityopts = ('Fes', 'Rabat','Casablanca')
        self.City = StringVar()
        self.City_label = Label(self.txt_frame, text="Ville", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=1, column=7, sticky=W, padx=5)
        self.Citycombo = ttk.Combobox(self.txt_frame, textvariable=self.City, values=self.Cityopts,  width=16,  font=("bold",12))
        self.Citycombo.grid(row=1,column=8) 
        
        #Department
        self.Department = StringVar()
        self.Department_Label = Label(self.txt_frame, text="Dépt", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=1, column=9, sticky=W, padx=5 )
        self.Department_entry =tk. Entry(self.txt_frame, textvariable=self.Department, width=17, bd=3, font=("bold",12))
        self.Department_entry.grid(row=1,column=10)
       
        #Type de contrat
        
        self.opts = ('CDI', 'CDD', 'OTHER')
        self.contrat = StringVar()
        self.contrat_label = Label(self.txt_frame, text="Contrat", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=4, column=3, sticky=W, padx=5)
        self.combof = ttk.Combobox(self.txt_frame, textvariable=self.contrat, values=self.opts,  width=16,  font=("bold",12))
        self.combof.grid(row=4,column=4) 
        
        #Last project name
        self.project = StringVar()
        self.project_Label = Label(self.txt_frame, text="Projet", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=3, column=0, sticky=W, padx=10 )
        self.project_entry = tk.Entry(self.txt_frame, textvariable=self.project , width=17, bd=3, font=("bold",12))
        self.project_entry.grid(row=3,column=1)
        
        #Last position
        self.position = StringVar()
        self.position_Label = Label(self.txt_frame, text="Position", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=3, column=3, sticky=W, padx=5 )
        self.position_entry = tk.Entry(self.txt_frame,textvariable=self.position, width=17, bd=3, font=("bold",12))
        self.position_entry.grid(row=3,column=4)
        
        #Hiring date cal2=Calendar(win,selectmode='day', year=2022, month=5, day=22)      

        self.dates = StringVar()
        self.dates_Label = Label(self.txt_frame, text="Date d'entrée", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=3, column=5, sticky=W, padx=5 )
        self.dates_entry = DateEntry(self.txt_frame,selectmode='day',width=16, bd=3, font=("bold",12),textvariable=self.dates,date_patern="yyy-mm-dd")
        self.dates_entry.grid(row=3,column=6)
                
        # Level
        self.opts = ('IEJ', 'IE2', 'SP1', 'TJ')
        self.level = StringVar()
        self.level_label = Label(self.txt_frame, text="Niveau", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=3, column=7, sticky=W, padx=5)
        self.combo_lev = ttk.Combobox(self.txt_frame, textvariable=self.level, values=self.opts,  width=16,  font=("bold",12))
        
        self.combo_lev.grid(row=3,column=8) 
        
        
        # PCI
        self.pci = StringVar()
        self.pci_Label = Label(self.txt_frame, text="€", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=3, column=9, sticky=W, padx=5 )
        self.pci_entry = tk.Entry(self.txt_frame, textvariable=self.pci, width=17, bd=3, font=("bold",12))
        self.pci_entry.grid(row=3,column=10)
        
        #Diplome
        self.opts = ('BAC+2','BAC+3','BAC+4', 'BAC+5')
        self.bac = StringVar()
        self.bac = Label(self.txt_frame, text="Diplôme", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=4, column=5, sticky=W, padx=5)
        self.bac_lev = ttk.Combobox(self.txt_frame, textvariable=self.bac, values=self.opts,  width=16,  font=("bold",12))
        
        self.bac_lev.grid(row=4,column=6) 
        
        # Specific position
        self.Specific_position = StringVar()
        self.Specific_position_Label = Label(self.txt_frame, text="Position spécifique", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=4, column=0, sticky=W, padx=10 )
        self.Specific_position_entry = tk.Entry(self.txt_frame, textvariable=self.Specific_position, width=17, bd=3, font=("bold",12))
        self.Specific_position_entry.grid(row=4,column=1)
        
        # Performance
        self.Performance = StringVar()
        self.Performance_Label = Label(self.txt_frame, text="Performance", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=4, column=7, sticky=W, padx=5 )
        self.Performance_entry = Spinbox(self.txt_frame, textvariable=self.Performance, width=16, bd=3, font=("bold",12), from_=1 , to=5)
        self.Performance_entry.grid(row=4,column=8)
        
        # Potential
        self.optsPotential = ('X1','X2','X3','X4')
        self.Potential = StringVar()
        self.Potential_label = Label(self.txt_frame, text="Potentiel", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=4, column=9, sticky=W, padx=5)
        self.comboPotential = ttk.Combobox(self.txt_frame, textvariable=self.Potential, values=self.optsPotential,  width=16,  font=("bold",12))
        self.comboPotential.grid(row=4,column=10) 
        # Studies
        self.Studies = StringVar()
        self.Studies_Label = Label(self.txt_frame, text="Etudes", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=5, column=0, sticky=W, padx=10 )
        self.Studies_entry = tk.Entry(self.txt_frame, textvariable=self.Studies, width=17, bd=3, font=("bold",12))
        self.Studies_entry.grid(row=5,column=1)
        
        # Scoring
        self.optsScoring = ('Faible','Insatisfaisant','Satisfaisant','Très satisfaisant')
        self.Scoring = StringVar()
        self.Scoring_label = Label(self.txt_frame, text="Scoring", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=5, column=3, sticky=W, padx=5)
        self.comboScoring= ttk.Combobox(self.txt_frame, textvariable=self.Scoring, values=self.optsScoring,  width=16,  font=("bold",12))
        self.comboScoring.grid(row=5,column=4)
        
        # Principal competency
        self.Principal_competency = StringVar()
        self.Principal_competency_Label = Label(self.txt_frame, text="Compétence principale", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=5, column=5, sticky=W, padx=5 )
        self.Principal_competency_entry = tk.Entry(self.txt_frame, textvariable=self.Principal_competency, width=17, bd=3, font=("bold",12))
        self.Principal_competency_entry.grid(row=5,column=6)
        
        # Secondary competency
        self.Secondary_competency = StringVar()
        self.Secondary_competency_Label = Label(self.txt_frame, text="Compétence secondaire", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=5, column=7, sticky=W, padx=5 )
        self.Secondary_competency_entry = tk.Entry(self.txt_frame, textvariable=self.Secondary_competency, width=17, bd=3, font=("bold",12))
        self.Secondary_competency_entry.grid(row=5,column=8)
        
        # Hiring date SISE
        self.datesSISE = StringVar()
        self.datesSISE_Label = Label(self.txt_frame, text="Date d'entrée au SISE", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=5, column=9, sticky=W, padx=5 )
        self.datesSISE_entry = DateEntry(self.txt_frame,selectmode='day',width=16, bd=3, font=("bold",12),date_patern="yyy-mm-dd",textvariable=self.datesSISE)
        self.datesSISE_entry.grid(row=5,column=10)
        
        #MF
        self.optsMF = ('male','Female')
        self.MF = StringVar()
        self.MF_label = Label(self.txt_frame, text="M / F", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=6, column=0, sticky=W, padx=10)
        self.comboMF= ttk.Combobox(self.txt_frame, textvariable=self.MF, values=self.optsMF,  width=16,  font=("bold",12))
        self.comboMF.grid(row=6,column=1)
        
        #E/T
        self.optsET = ('Engineers','Technicians')
        self.ET = StringVar()
        self.ET_label = Label(self.txt_frame, text="E / T", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=6, column=3, sticky=W, padx=10)
        self.comboET= ttk.Combobox(self.txt_frame, textvariable=self.ET, values=self.optsET,  width=16,  font=("bold",12))
        self.comboET.grid(row=6,column=4)
        
        # Leaving Reasons
        self.Leaving_Reasons = StringVar()
        self.Leaving_Reasons_Label = Label(self.txt_frame, text="Raison de départ", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=6, column=5, sticky=W, padx=5 )
        self.Leaving_Reasons_entry = tk.Entry(self.txt_frame, textvariable=self.Leaving_Reasons, width=17, bd=3, font=("bold",12))
        self.Leaving_Reasons_entry.grid(row=6,column=6)
        
        # Absenteeism Reasons
        self.Absenteeism_Reasons = StringVar()
        self.Absenteeism_Reasons_Label = Label(self.txt_frame, text="Raison d'abscence", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=6, column=7, sticky=W, padx=5 )
        self.Absenteeism_Reasons_entry = tk.Entry(self.txt_frame, textvariable=self.Absenteeism_Reasons, width=17, bd=3, font=("bold",12))
        self.Absenteeism_Reasons_entry.grid(row=6,column=8)
        
        # Previous exp
        self.Previous_exp = StringVar()
        self.Previous_exp_Label = Label(self.txt_frame, text="Expérience précédente", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=6, column=9, sticky=W, padx=5 )
        self.Previous_exp_entry = Spinbox(self.txt_frame, textvariable=self.Previous_exp, width=16, bd=3, font=("bold",12), from_=0 , to=30)
        self.Previous_exp_entry.grid(row=6,column=10)
        
        # Mission
        self.Mission = StringVar()
        self.Mission_Label = Label(self.txt_frame, text="Mission", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=7, column=0, sticky=W, padx=10)
        self.Mission_entry = tk.Entry(self.txt_frame, textvariable=self.Mission, width=17, bd=3, font=("bold",12))
        self.Mission_entry.grid(row=7,column=1)
        
        # DATE CDI
        self.datesCDI = StringVar()
        self.datesCDI_Label = Label(self.txt_frame, text="DATE CDI", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=7, column=3, sticky=W, padx=5 )
        self.datesCDI_entry = DateEntry(self.txt_frame,selectmode='day',width=16, bd=3, font=("bold",12))
        self.datesCDI_entry.grid(row=7,column=4)
        
        # Exit Date
        self.datesExit_Date = StringVar()
        self.datesExit_Date_Label = Label(self.txt_frame, text="Date de sortie", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=7, column=5, sticky=W, padx=5 )
        self.datesExit_Date_entry = DateEntry(self.txt_frame,selectmode='day',width=16, bd=3, font=("bold",12))
        self.datesExit_Date_entry.grid(row=7,column=6)
        
        # Expérience ALTEN
        self.Expérience_ALTEN = StringVar()
        self.Expérience_ALTEN_Label = Label(self.txt_frame, text="Expérience ALTEN", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=7, column=7, sticky=W, padx=5 )
        self.Expérience_ALTEN_entry = Spinbox(self.txt_frame, textvariable=self.Expérience_ALTEN, width=16, bd=3, font=("bold",12), from_=0 , to=30)
        self.Expérience_ALTEN_entry.grid(row=7,column=8)
        
        # Expérience totale
        self.Expérience_totale = StringVar()
        self.Expérience_totale_Label = Label(self.txt_frame, text="Expérience totale", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=7, column=9, sticky=W, padx=5 )
        self.Expérience_totale_entry = Spinbox(self.txt_frame, textvariable=self.Expérience_totale, width=16, bd=3, font=("bold",12), from_=0 , to=30)
        self.Expérience_totale_entry.grid(row=7,column=10)
        
        # Monthly salary local
        self.Monthly_salary_local = StringVar()
        self.Monthly_salary_local_Label = Label(self.txt_frame, text="Salaire mensuel", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=8, column=0, sticky=W, padx=10)
        self.Monthly_salary_local_entry = tk.Entry(self.txt_frame, textvariable=self.Monthly_salary_local, width=17, bd=3, font=("bold",12))
        self.Monthly_salary_local_entry.grid(row=8,column=1)
        

        #Prenom
        self.Prenom = StringVar()
        self.Prenom_Label = Label(self.txt_frame, text="Prenom", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=8, column=3, sticky=W, padx=5)
        self.Prenom_entry = tk.Entry(self.txt_frame, textvariable=self.Prenom, width=17, bd=3, font=("bold",12))
        self.Prenom_entry.grid(row=8,column=4)
        
        # Nom
        self.Nom = StringVar()
        self.Nom_Label = Label(self.txt_frame, text="Nom", font=("bold", 13), bg="#CBDEF1", fg="black").grid(row=8, column=5, sticky=W, padx=5)
        self.Nom_entry = tk.Entry(self.txt_frame, textvariable=self.Nom, width=17, bd=3, font=("bold",12))
        self.Nom_entry.grid(row=8,column=6)
        
        
        #treeview
        self.list_frame = Frame(self.detail_frame, bd=2, relief=RIDGE , bg="white")
        self.list_frame.place(x=0, y=0,  width=1620, height=468)
         #scoller treeview
        scroll_x = Scrollbar(self.list_frame, orient=HORIZONTAL)
        scroll_y = Scrollbar(self.list_frame, orient=VERTICAL)
        
        design = ttk.Style()
        design.theme_use("clam")
       
        self.data_list = ttk.Treeview(self.list_frame  , height=12, column=("ID", "Manager", "Company", "City", "Department", "Type of contract",
       "Last project name", "Last Position", "Hiring date", "Level", "€",
       "Specific position", "Diplôme", "Performance", "Potential", "Studies",
       "Scoring", "Principal competency", "Secondary competency",
       "Hiring date SISE", "M/F", "E/T", "Leaving Reasons",
       "Absenteeism Reasons", "Previous exp", "Mission", "DATE CDI",
       "Exit Date", "Expérience ALTEN", "Expérience totale",
       "Monthly salary local", "Prenom", "Nom"),
                                      xscrollcommand = scroll_x.set , yscrollcommand= scroll_y.set)
        
        scroll_x.pack(side=BOTTOM , fill=X)
        scroll_y.pack(side=RIGHT , fill=Y)
           
        self.data_list.configure(yscrollcommand= scroll_x.set)
        scroll_x.configure(command = self.data_list.xview)
        
        self.data_list.configure(yscrollcommand= scroll_y.set)
        scroll_y.configure(command = self.data_list.yview)
        
        self.data_list.heading("ID",text="ID")
        self.data_list.heading("Manager",text="Manager")
        self.data_list.heading("Company",text="Company")
        self.data_list.heading( "City",text= "City")
        self.data_list.heading("Department",text="Department")
        self.data_list.heading("Type of contract",text="Type of contract")
        self.data_list.heading("Last project name",text="Last project name")
        self.data_list.heading("Last Position",text="Last Position")
        self.data_list.heading("Hiring date",text="Hiring date")
        self.data_list.heading("Level",text="Level")
        self.data_list.heading("€",text="€")
        self.data_list.heading("Specific position",text="Specific position")
        self.data_list.heading( "Diplôme",text="Diplôme")
        self.data_list.heading( "Performance",text= "Performance")
        self.data_list.heading("Potential",text="Potential")
        self.data_list.heading("Studies",text="Studies")
        self.data_list.heading("Scoring",text="Scoring")
        self.data_list.heading("Principal competency",text="Principal competency")
        self.data_list.heading("Secondary competency",text="Secondary competency")
        self.data_list.heading("Hiring date SISE",text="Hiring date SISE")
        self.data_list.heading("M/F",text="M/F")
        self.data_list.heading("E/T",text="E/T")
        self.data_list.heading("Leaving Reasons",text="Leaving Reasons")
        self.data_list.heading("Absenteeism Reasons",text="Absenteeism Reasons")
        self.data_list.heading("Previous exp",text="Previous exp")
        self.data_list.heading("Mission",text="Mission")
        self.data_list.heading( "DATE CDI",text="DATE CDI")
        self.data_list.heading("Exit Date",text="Exit Date")
        self.data_list.heading("Expérience ALTEN",text="Expérience ALTEN")
        self.data_list.heading("Expérience totale",text="Expérience totale")
        self.data_list.heading("Monthly salary local",text="Monthly salary local")
        self.data_list.heading("Prenom",text="Prenom")
        self.data_list.heading("Nom",text="Nom")
        
        self.data_list["show"] = "headings"
     
        self.data_list.column("ID", width=80, anchor=tk.CENTER)
        self.data_list.column("Manager", width=80, anchor=tk.CENTER)
        self.data_list.column("Company", width=80, anchor=tk.CENTER)
        self.data_list.column( "City", width=80, anchor=tk.CENTER)
        self.data_list.column("Department", width=80, anchor=tk.CENTER)
        self.data_list.column("Type of contract", width=80, anchor=tk.CENTER)
        self.data_list.column("Last project name", width=80, anchor=tk.CENTER)
        self.data_list.column("Last Position", width=80, anchor=tk.CENTER)
        self.data_list.column("Hiring date", width=80, anchor=tk.CENTER)
        self.data_list.column("Level", width=80, anchor=tk.CENTER)
        self.data_list.column("€", width=80, anchor=tk.CENTER)
        self.data_list.column("Specific position", width=80, anchor=tk.CENTER)
        self.data_list.column( "Diplôme", width=80, anchor=tk.CENTER)
        self.data_list.column( "Performance", width=80, anchor=tk.CENTER)
        self.data_list.column("Potential", width=80, anchor=tk.CENTER)
        self.data_list.column("Studies", width=80, anchor=tk.CENTER)
        self.data_list.column("Scoring", width=80, anchor=tk.CENTER)
        self.data_list.column("Principal competency", width=80, anchor=tk.CENTER)
        self.data_list.column("Secondary competency", width=80, anchor=tk.CENTER)
        self.data_list.column("Hiring date SISE", width=80, anchor=tk.CENTER)
        self.data_list.column("M/F", width=80, anchor=tk.CENTER)
        self.data_list.column("E/T", width=80, anchor=tk.CENTER)
        self.data_list.column("Leaving Reasons", width=80, anchor=tk.CENTER)
        self.data_list.column("Absenteeism Reasons", width=80, anchor=tk.CENTER)
        self.data_list.column("Previous exp", width=80, anchor=tk.CENTER)
        self.data_list.column("Mission", width=80, anchor=tk.CENTER)
        self.data_list.column( "DATE CDI", width=80, anchor=tk.CENTER)
        self.data_list.column("Exit Date", width=80, anchor=tk.CENTER)
        self.data_list.column("Expérience ALTEN", width=80, anchor=tk.CENTER)
        self.data_list.column("Expérience totale", width=80, anchor=tk.CENTER)
        self.data_list.column("Monthly salary local", width=80, anchor=tk.CENTER)
        self.data_list.column("Prenom", width=80, anchor=tk.CENTER)
        self.data_list.column("Nom",width=80, anchor=tk.CENTER)
        self.id = 0
        self.data_list.pack(fill=BOTH , expand=1)
        self.data_list.bind('<ButtonRealease-1>')

        
        
        df_rows = employee_list.to_numpy().tolist()

        for row in df_rows:
         self.data_list.insert("", "end", values=(row))
    
    def delete_data(self):
        
        select=self.data_list.focus()
       
        
        #values = self.data_list.item(select,"values")
        #print (values)
        #employee_list.drop(employee_list.iloc[self.data_list.index(select),0].index, inplace=True, axis=0)
        valeur_supp = employee_list.iloc[self.data_list.index(select),0]
        List_index = employee_list.loc[employee_list["ID"] == valeur_supp ].index
        self.data_list.delete(select)
        
        for i in List_index :
            employee_list.drop(i,inplace= True , axis=0)
            #employee_list.reset_index(inplace=True)
        employee_list.to_excel(wb_path, index=False)
            

        #print(employee_list.iloc[self.data_list.index(select),0])
        
        #print self.data_list.selection
        #employee_list.to_excel(wb_path)
        
    def load_workbook(wb_path):
        if os.path.exists(wb_path):
              return openpyxl.load_workbook(wb_path)
        else:
             return " File not found"
    def select_record(self):
        self.employe_entry.delete(0,END)
        self.comborr.delete(0,END)
        self.Company_entry.delete(0,END)
        self.Citycombo.delete(0,END)
        self.Department_entry.delete(0,END)
        self.combof.delete(0,END)
        self.project_entry.delete(0,END)
        self.position_entry.delete(0,END)
        self.dates_entry.delete(0,END)
        self.combo_lev.delete(0,END)
        self.pci_entry.delete(0,END)
        self.Specific_position_entry.delete(0,END)
        self.bac_lev.delete(0,END)
        self.Performance_entry.delete(0,END)
        self.comboPotential.delete(0,END)
        self.Studies_entry.delete(0,END)
        self.comboScoring.delete(0,END)
        self.Principal_competency_entry.delete(0,END)
        self.Secondary_competency_entry.delete(0,END) 
        self.datesSISE_entry.delete(0,END)
        self.comboMF.delete(0,END)
        self.comboET.delete(0,END) 
        self.Leaving_Reasons_entry.delete(0,END)
        self.Absenteeism_Reasons_entry.delete(0,END)
        self.Previous_exp_entry.delete(0,END)
        self.Mission_entry.delete(0,END)
        self.datesCDI_entry.delete(0,END)
        self.datesExit_Date_entry.delete(0,END)
        self.Expérience_ALTEN_entry.delete(0,END)
        self.Expérience_totale_entry.delete(0,END)
        self.Monthly_salary_local_entry.delete(0,END)
        self.Prenom_entry.delete(0,END)
        self.Nom_entry.delete(0,END)
        
        select=self.data_list.focus()
    #save new data 
        
        
        values = self.data_list.item(select,"values")
        
        self.employe_entry.insert(0,values[0])
        self.comborr.insert(0,values[1])
        self.Company_entry.insert(0,values[2])
        self.Citycombo.insert(0,values[3])
        self.Department_entry.insert(0,values[4])
        self.combof.insert(0,values[5])
        self.project_entry.insert(0,values[6]) 
        self.position_entry.insert(0,values[7])
        self.dates_entry.insert(0,values[8])
        self.combo_lev.insert(0,values[9])
        self.pci_entry.insert(0,values[10]) 
        self.Specific_position_entry.insert(0,values[11])
        self.bac_lev.insert(0,values[12])
        self.Performance_entry.insert(0,values[13])
        self.comboPotential.insert(0,values[14])
        self.Studies_entry.insert(0,values[15])
        self.comboScoring.insert(0,values[16])
        self.Principal_competency_entry.insert(0,values[17])
        self.Secondary_competency_entry.insert(0,values[18]) 
        self.datesSISE_entry.insert(0,values[19])
        self.comboMF.insert(0,values[20])
        self.comboET.insert(0,values[21])
        self.Leaving_Reasons_entry.insert(0,values[22])
        self.Absenteeism_Reasons_entry.insert(0,values[23])
        self.Previous_exp_entry.insert(0,values[24])
        self.Mission_entry.insert(0,values[25])
        self.datesCDI_entry.insert(0,values[26]),
        self.datesExit_Date_entry.insert(0,values[27]),
        self.Expérience_ALTEN_entry.insert(0,values[28]) 
        self.Expérience_totale_entry.insert(0,values[29]) 
        self.Monthly_salary_local_entry.insert(0,values[30])
        self.Prenom_entry.insert(0,values[31])
        self.Nom_entry.insert(0,values[32])
    def clear(self):
        self.employe_entry.delete(0,END)
        self.comborr.delete(0,END)
        self.Company_entry.delete(0,END)
        self.Citycombo.delete(0,END)
        self.Department_entry.delete(0,END)
        self.combof.delete(0,END)
        self.project_entry.delete(0,END)
        self.position_entry.delete(0,END)
        self.dates_entry.delete(0,END)
        self.combo_lev.delete(0,END)
        self.pci_entry.delete(0,END)
        self.Specific_position_entry.delete(0,END)
        self.bac_lev.delete(0,END)
        self.Performance_entry.delete(0,END)
        self.comboPotential.delete(0,END)
        self.Studies_entry.delete(0,END)
        self.comboScoring.delete(0,END)
        self.Principal_competency_entry.delete(0,END)
        self.Secondary_competency_entry.delete(0,END) 
        self.datesSISE_entry.delete(0,END)
        self.comboMF.delete(0,END)
        self.comboET.delete(0,END) 
        self.Leaving_Reasons_entry.delete(0,END)
        self.Absenteeism_Reasons_entry.delete(0,END)
        self.Previous_exp_entry.delete(0,END)
        self.Mission_entry.delete(0,END)
        self.datesCDI_entry.delete(0,END)
        self.datesExit_Date_entry.delete(0,END)
        self.Expérience_ALTEN_entry.delete(0,END)
        self.Expérience_totale_entry.delete(0,END)
        self.Monthly_salary_local_entry.delete(0,END)
        self.Prenom_entry.delete(0,END)
        self.Nom_entry.delete(0,END)
        
        
        
    def Update(self):

        select=self.data_list.focus()
        self.data_list.item(select,text = "",values=(self.employe_entry.get(), self.comborr.get(), self.Company_entry.get(),
                                                  self.Citycombo.get(),self.Department_entry.get(),
                                                  self.combof.get(),self.project_entry.get(), self.position_entry.get(),
                                                  self.dates_entry.get_date(),self.combo_lev.get(),
                                                  self.pci_entry.get(), self.Specific_position_entry.get(),
                                                  self.bac_lev.get(),self.Performance_entry.get(),self.comboPotential.get(),
                                                  self.Studies_entry.get(),self.comboScoring.get(),self.Principal_competency_entry.get(),
                                                  self.Secondary_competency_entry.get(), self.datesSISE_entry.get(),self.comboMF.get(),
                                                  self.comboET.get(), self.Leaving_Reasons_entry.get(),self.Absenteeism_Reasons_entry.get(),
                                                  self.Previous_exp_entry.get(),self.Mission_entry.get(),self.datesCDI_entry.get(),self.datesExit_Date_entry.get(),
                                                  self.Expérience_ALTEN_entry.get(), self.Expérience_totale_entry.get(), self.Monthly_salary_local_entry.get(),self.Prenom_entry.get(),self.Nom_entry.get()))
        
        employee_list.iloc[self.data_list.index(self.data_list.selection()),1]=self.comborr.get()     
        employee_list.iloc[self.data_list.index(self.data_list.selection()),2]=self.Company_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),3]=self.Citycombo.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),4]=self.Department_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),5]=self.combof.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),6]=self.project_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),7]=self.position_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),8]=self.dates_entry.get_date()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),9]=self.combo_lev.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),10]=self.pci_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),11]=self.Specific_position_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),12]=self.bac_lev.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),13]=self.Performance_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),14]=self.comboPotential.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),15]=self.Studies_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),16]=self.comboScoring.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),17]=self.Principal_competency_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),18]=self.Secondary_competency_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),19]=self.datesSISE_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),20]=self.comboMF.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),21]=self.comboET.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),22]=self.Leaving_Reasons_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),23]=self.Absenteeism_Reasons_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),24]=self.Previous_exp_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),25]=self.Mission_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),26]=self.datesCDI_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),27]=self.datesExit_Date_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),28]=self.Expérience_ALTEN_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),29]=self.Expérience_totale_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),30]=self.Monthly_salary_local_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),31]=self.Prenom_entry.get()
        employee_list.iloc[self.data_list.index(self.data_list.selection()),31]=self.Nom_entry.get()
        #employee_list.reset_index(inplace=False)
        employee_list.to_excel(wb_path, index=False)    

        
        
        
        
        self.employe_entry.delete(0,END)
        self.comborr.delete(0,END)
        self.Company_entry.delete(0,END)
        self.Citycombo.delete(0,END)
        self.Department_entry.delete(0,END)
        self.combof.delete(0,END)
        self.project_entry.delete(0,END)
        self.position_entry.delete(0,END)
        self.dates_entry.delete(0,END)
        self.combo_lev.delete(0,END)
        self.pci_entry.delete(0,END)
        self.Specific_position_entry.delete(0,END)
        self.bac_lev.delete(0,END)
        self.Performance_entry.delete(0,END)
        self.comboPotential.delete(0,END)
        self.Studies_entry.delete(0,END)
        self.comboScoring.delete(0,END)
        self.Principal_competency_entry.delete(0,END)
        self.Secondary_competency_entry.delete(0,END) 
        self.datesSISE_entry.delete(0,END)
        self.comboMF.delete(0,END)
        self.comboET.delete(0,END) 
        self.Leaving_Reasons_entry.delete(0,END)
        self.Absenteeism_Reasons_entry.delete(0,END)
        self.Previous_exp_entry.delete(0,END)
        self.Mission_entry.delete(0,END)
        self.datesCDI_entry.delete(0,END)
        self.datesExit_Date_entry.delete(0,END)
        self.Expérience_ALTEN_entry.delete(0,END)
        self.Expérience_totale_entry.delete(0,END)
        self.Monthly_salary_local_entry.delete(0,END)
        self.Prenom_entry.delete(0,END)
        self.Nom_entry.delete(0,END)
          
             
root = Tk()
obj = studentForm(root)
root.deiconify()
image1 = Image.open(r'C:\Users\Administrateur\Downloads\Alten.png')
resized_image= image1.resize((200,72), Image.ANTIALIAS)
test = ImageTk.PhotoImage(resized_image)

label1 = tk.Label(image=test)
label1.image = test
label1.place(x=0, y=0)
root.mainloop()
        
        
 
        
        
        
        
        
        